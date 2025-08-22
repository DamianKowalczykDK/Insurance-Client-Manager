from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from src.excel.type.style_type import CellStyle, apply_style
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook, Workbook
from typing import Callable, Mapping
from pathlib import Path


class ExcelManager:
    """Manager for handling Excel workbook operations such as adding rows,
    formatting, styling, and column adjustments.

    Attributes:
        filepath (str): Path to the Excel file.
        sheet_name (str): Default worksheet name.
        workbook (Workbook): OpenPyXL workbook instance.
    """

    def __init__(
            self,
            filepath: str,
            sheet_name: str = "Clients",
    ) -> None:
        """Initialize an ExcelManager.

        Args:
            filepath (str): Path to the Excel file.
            sheet_name (str, optional): Name of the worksheet. Defaults to "Clients".
        """
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.workbook = self._load_or_create()

        if self.sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(self.sheet_name)

    def get_sheet(self, name: str | None = None) -> Worksheet:
        """Get a worksheet by name.

        Args:
            name (str | None, optional): Worksheet name. Defaults to the default sheet.

        Returns:
            Worksheet: The requested worksheet.
        """
        return self.workbook[name or self.sheet_name]

    def save(self) -> None:
        """Save the workbook to the file path."""
        self.workbook.save(self.filepath)

    # ------------------------------------------------------------------------------------------------------------------
    #  Data
    # ------------------------------------------------------------------------------------------------------------------

    def add_row[T: Mapping](self, sheet_name: str, data: T, row_idx: int = 1, col_letter: str = 'A') -> None:
        """Insert a row of data into the worksheet.

        Args:
            sheet_name (str): Name of the worksheet.
            data (Mapping): Data to insert as row values.
            row_idx (int, optional): Target row index. Defaults to 1.
            col_letter (str, optional): Starting column letter. Defaults to 'A'.
        """
        ws = self.get_sheet(sheet_name)
        start_col_idx = column_index_from_string(col_letter)

        for offset, val in enumerate(data.values()):
            ws.cell(row=row_idx, column=start_col_idx + offset).value = val

    def apply_str_conversion_for_ranges(
            self,
            converter_fn: Callable[[str], str],
            uppercase_cell_ranges: list[str] | None = None
    ) -> None:
        """Apply a string conversion function to cell values in given ranges.

        Args:
            converter_fn (Callable[[str], str]): Function to convert cell values.
            uppercase_cell_ranges (list[str] | None, optional): List of cell ranges. Defaults to None.
        """
        for cell_range in uppercase_cell_ranges or []:
            ws = self.get_sheet()
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)

            save_min_row = min_row if min_row is not None else ws.min_row
            save_max_row = max_row if max_row is not None else ws.max_row

            for row in ws.iter_rows(
                min_row=max(ws.min_row, save_min_row),
                max_row=min(ws.max_row, save_max_row),
                min_col=min_col,
                max_col=max_col):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = converter_fn(cell.value)

    # ------------------------------------------------------------------------------------------------------------------
    #  Autofit
    # ------------------------------------------------------------------------------------------------------------------

    def autofit_column_widths(self, sheet_name: str | None = None, offset_dim: int = 5) -> None:
        """Automatically adjust column widths based on content length.

        Args:
            sheet_name (str | None, optional): Worksheet name. Defaults to the default sheet.
            offset_dim (int, optional): Additional padding for width. Defaults to 5.
        """
        ws: Worksheet = self.get_sheet(sheet_name)

        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
            col_idx = col_cells[0].column
            max_length = 0

            for cell in col_cells:
                value = cell.value
                if value is not None:
                    try:
                        max_length = max(max_length, len(str(value)))
                    except Exception:
                        continue
            if col_idx:
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max_length + offset_dim

    def get_last_row_in_col(self, col_letter: str, sheet_name: str | None = None) -> int:
        """Get the index of the last non-empty row in a column.

        Args:
            col_letter (str): Target column letter.
            sheet_name (str | None, optional): Worksheet name. Defaults to the default sheet.

        Returns:
            int: Last non-empty row index.
        """
        ws = self.get_sheet(sheet_name)
        for row in range(ws.max_row, 0, -1):
            if ws[f"{col_letter}{row}"].value not in (None, ""):
                return row
        return 1

    # -----------------------------------------------------------------------------------------------------
    # Style
    # -----------------------------------------------------------------------------------------------------

    def style_cell(self, cell_ref: str, style: CellStyle, sheet_name: str | None = None) -> None:
        """Apply a style to a single cell.

        Args:
            cell_ref (str): Cell reference (e.g., 'A1').
            style (CellStyle): Style object to apply.
            sheet_name (str | None, optional): Worksheet name. Defaults to the default sheet.
        """
        ws = self.get_sheet(sheet_name)
        cell = ws[cell_ref]
        apply_style(cell, style)

    def style_table_area(
            self,
            start_col_letter: str,
            headers: list[str],
            header_style: CellStyle | None = None,
            row_style: CellStyle | None = None,
            sheet_name: str | None = None,
    ) -> None:
        """Apply styles to a table area, including headers and rows.

        Args:
            start_col_letter (str): Starting column letter for the table.
            headers (list[str]): List of header names.
            header_style (CellStyle | None, optional): Style for headers. Defaults to None.
            row_style (CellStyle | None, optional): Style for rows. Defaults to None.
            sheet_name (str | None, optional): Worksheet name. Defaults to the default sheet.
        """
        start_idx = column_index_from_string(start_col_letter)
        max_row = self.get_last_row_in_col(start_col_letter, sheet_name)

        for offset in range(len(headers)):
            col_letter = get_column_letter(start_idx + offset)

            if header_style:
               self.style_cell(f"{col_letter}1", header_style, sheet_name)

            if row_style and max_row >= 2:
                for row in range(2, max_row + 1):
                    self.style_cell(f"{col_letter}{row}", row_style, sheet_name)

    # -----------------------------------------------------------------------------------------------------
    # Format
    # -----------------------------------------------------------------------------------------------------

    def set_column_format(
            self,
            col_letter: str,
            data_format: str,
            start_row: int = 1,
            end_row: int | None = None,
            sheet_name: str | None = None
    ) -> None:
        """Set number format for a column range.

        Args:
            col_letter (str): Target column letter.
            data_format (str): Format string (e.g., '0.00').
            start_row (int, optional): Starting row index. Defaults to 1.
            end_row (int | None, optional): Ending row index. Defaults to the last row.
            sheet_name (str | None, optional): Worksheet name. Defaults to the default sheet.
        """
        ws = self.get_sheet(sheet_name)
        end_row = end_row or ws.max_row
        col_idx = column_index_from_string(col_letter)

        for row_idx in range(start_row, end_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = data_format

    # -----------------------------------------------------------------------------------------------------
    # Method auxiliary
    # -----------------------------------------------------------------------------------------------------

    def _load_or_create(self) -> Workbook:
        """Load an existing workbook or create a new one.

        Returns:
            Workbook: OpenPyXL workbook instance.
        """
        path = Path(self.filepath)
        if path.exists():
            return load_workbook(self.filepath)
        else:
            workbook = Workbook()
            if workbook.active is not None:
                workbook.remove(workbook.active)
            return workbook
