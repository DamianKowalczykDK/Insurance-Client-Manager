from calendar import month

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Alignment
from src.excel.manager.base_manager import ExcelManager
from openpyxl.worksheet.worksheet import Worksheet
from src.excel.type.style_type import CellStyle
from datetime import datetime, date, timedelta
from src.model.client import ClientDict
from typing import override, cast


class ClientExcelManager(ExcelManager):
    """Manages client-related data in an Excel file.

    This class extends `ExcelManager` to provide functionality for handling
    client records, insurance companies, and summary tables.
    It supports inserting, updating, deleting, and loading client data,
    as well as generating summary reports and applying styling.
    """

    def __init__(
        self,
        filepath: str,
        sheet_name: str = "Clients",

        main_table_headers: list[str] | None = None,
        company_table_headers: list[str] | None = None,
        summary_table_headers: list[str] | None = None,

        header_style: CellStyle | None = None,
        row_style: CellStyle | None = None,
        overdue_style: CellStyle | None = None,

        main_table_start_col: str = "A",
        company_table_start_col: str = "I",

        uppercase_columns: list[str] | None = None,
        ratio: float = 0.74,
    ) -> None:
        """Initialize the client Excel manager.

        Args:
            filepath: Path to the Excel file.
            sheet_name: Worksheet name to manage.
            main_table_headers: Headers for the main client table.
            company_table_headers: Headers for the company summary table.
            summary_table_headers: Headers for the summary metrics table.
            header_style: Style for table headers.
            row_style: Style for table rows.
            overdue_style: Style for overdue payments.
            main_table_start_col: Starting column for the main table.
            company_table_start_col: Starting column for the company table.
            uppercase_columns: List of column letters to apply uppercase conversion.
            ratio: Ratio used for calculating metrics in the summary table.
        """
        super().__init__(filepath, sheet_name)
        self.main_table_headers = (main_table_headers or
                                   ["NAME", "EMAIL", "INSURANCE_COMPANY", "CAR_MODEL", "CAR_YEAR", "PRICE", "NEXT_PAYMENT"])
        self.company_table_headers = company_table_headers or ["INSURANCE_COMPANY", "CLIENT"]
        self.summary_table_headers = summary_table_headers or ["METRIC", "VALUE"]
        self.ratio = ratio
        self.header_style = header_style
        self.row_style = row_style
        self.uppercase_columns = uppercase_columns or []
        self.overdue_style = overdue_style
        self.main_table_start_col = main_table_start_col
        self.company_table_start_col = company_table_start_col

        self._validate_headers()
        self.summary_table_start_col = self._validate_column_ranges()

        ws = self.get_sheet()
        if all(cell.value is None for cell in ws[1]):
            for col_idx, header in enumerate(self.main_table_headers, start=1):
                ws.cell(row=1, column=col_idx).value = header.replace("_", " ")
        self.update_summary_tables()

    # ------------------------------------------------------------------------------------------------------------------
    #  Data
    # ------------------------------------------------------------------------------------------------------------------

    def get_next_main_table_row(self) -> int:
        """Find the next available row in the main client table.

        Returns:
            int: Row index where the next client can be inserted.
        """
        ws = self.get_sheet()
        for row_idx in range(2, ws.max_row + 1):
            if not ws.cell(row=row_idx, column=1).value:
                return row_idx
        return ws.max_row + 1

    def insert_main_row(self, data: ClientDict) -> None:
        """Insert a new client row into the main table.

        Args:
            data: Dictionary containing client data.
        """
        insert_row = self.get_next_main_table_row()
        self.add_row(data=data, row_idx=insert_row, sheet_name=self.sheet_name)
        self.update_summary_tables()

    # ------------------------------------------------------------------------------------------------------------------
    #  Tables
    # ------------------------------------------------------------------------------------------------------------------

    def update_summary_tables(self) -> None:
        """Update all summary tables (companies, metrics)."""
        ws = self.get_sheet()
        company = self._extract_unique_insurance_company(ws)

        self._aplay_uppercase()

        self._update_simple_summary(
            ws,
            company,
            tuple(self.company_table_headers),
            self.company_table_start_col,
            "C"
        )

        self._update_metric_table(ws)
        self.save()

    def update_client_row(self, col_value: int, value: str, data: ClientDict) -> bool:
        """Update an existing client row based on a column value.

        Args:
            col_value: Index of the column to search in.
            value: Value to match in the given column.
            data: New client data to overwrite the row.

        Returns:
            bool: True if the row was updated, False otherwise.
        """
        ws = self.get_sheet()
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=col_value).value == value:
                for col_idx, v in enumerate(data.values(), start=1):
                    ws.cell(row=row_idx, column=col_idx).value = cast(str | int, v)
                self.update_summary_tables()
                return True
        return False

    def shift_payment_date(self, col_value: int, value: str, payment_date_col: int, days: int = 360) -> bool:
        """Shift a client's payment date by a given number of days.

        Args:
            col_value: Column index used to find the client.
            value: Value to match in the column.
            payment_date_col: Column index of the payment date.
            days: Number of days to shift the payment date by.

        Returns:
            bool: True if the date was updated, False otherwise.
        """
        ws = self.get_sheet()
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=col_value).value == value:
                data = ws.cell(row=row_idx, column=payment_date_col)
                v = data.value
                if isinstance(v, str):
                    current_date = datetime.strptime(v, "%Y-%m-%d").date()
                elif isinstance(v, date):
                    current_date = v
                else:
                    return False
                new_date = current_date + timedelta(days=days)
                data.value = new_date.strftime("%Y-%m-%d")
                self.save()
                return True
        return False

    def remove_client_row(self, col_value: int, value: str) -> bool:
        """Remove a client row based on a column value.

        Args:
            col_value: Column index to search.
            value: Value to match in the column.

        Returns:
            bool: True if a row was removed, False otherwise.
        """
        ws = self.get_sheet()
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=col_value).value == value:
                ws.delete_rows(row_idx)
                self.update_summary_tables()
                return True
        return False

    def load_client_row(self) -> list[ClientDict]:
        """Load all clients from the worksheet.

        Returns:
            list[ClientDict]: List of clients as dictionaries.
        """
        ws = self.get_sheet()
        clients_keys = list(ClientDict.__annotations__.keys())
        required_length = len(clients_keys)

        clients: list[ClientDict] = []
        for row in ws.iter_rows(2, values_only=True):
            try:
                if not row or any(val is None for val in row[:required_length]):
                    continue

                clients.append({
                    "name": str(row[0]),
                    "email": str(row[1]),
                    "insurance_company": str(row[2]),
                    "car_model": str(row[3]),
                    "car_year": int(str(row[4])),
                    "price": int(str(row[5])),
                    "next_payment": str(row[6]),
                })
            except Exception:
                continue
        return clients

    def overwrite_clients(self, clients: list[ClientDict]) -> None:
        """Overwrite all clients in the worksheet with new data.

        Args:
            clients: List of client dictionaries.
        """
        ws = self.get_sheet()
        ws.delete_rows(2, ws.max_row)

        for row_idx, client in enumerate(clients, start=2):
            self.add_row(sheet_name=self.sheet_name, data=client, row_idx=row_idx)

        self.update_summary_tables()

    @override
    def save(self) -> None:
        """Apply styles and save the Excel file."""
        self.style_table_area(self.main_table_start_col, self.main_table_headers, self.header_style, self.row_style)
        self.style_table_area(self.company_table_start_col, self.company_table_headers, self.header_style, self.row_style)

        self._style_summary_table()
        self._highlight_overdue_payment()

        super().autofit_column_widths()
        super().save()

    # -----------------------------------------------------------------------------------------------------
    # Method auxiliary
    # -----------------------------------------------------------------------------------------------------

    def _highlight_overdue_payment(self) -> None:
        """Highlight overdue payments in the main table using overdue style."""
        ws = self.get_sheet()
        today = datetime.today().date()
        col_idx = self.main_table_headers.index("NEXT_PAYMENT") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.value

            if not val:
                continue

            try:
                cell_date = (
                    datetime.strptime(val, "%Y-%m-%d").date()
                    if isinstance(val, str) else
                    val if isinstance(val, date) else
                    None
                )

                if cell_date and cell_date <= today and self.overdue_style is not None:
                    self.style_cell(cell.coordinate, self.overdue_style)
            except Exception:
                continue

    def _extract_unique_insurance_company(
        self,
        ws: Worksheet,
        company_column_name: str = "INSURANCE_COMPANY",
    ) -> list[str]:
        """Extract a list of unique insurance companies from the worksheet.

        Args:
            ws: Worksheet object.
            company_column_name: Column name where insurance companies are stored.

        Returns:
            list[str]: Sorted list of unique insurance companies.
        """
        company_idx = self.main_table_headers.index(company_column_name)
        insurance_company: set[str] = set()

        for row in ws.iter_rows(min_row=2, max_col=len(self.main_table_headers), values_only=True):
            company_val = row[company_idx] if company_idx < len(row) else None
            if company_val:
                insurance_company.add(str(company_val))

        return sorted(insurance_company)

    def _clear_column_range(self, ws: Worksheet, col_letter: str, start_row: int, end_row: int) -> None:
        """Clear values and styles from a given column range.

        Args:
            ws: Worksheet object.
            col_letter: Column letter.
            start_row: Start row index.
            end_row: End row index.
        """
        for row in range(start_row, end_row + 1):
            cell = ws[f"{col_letter}{row}"]
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = "General"

    def _update_simple_summary(
        self,
        ws: Worksheet,
        unique_values: list[str] | list[int],
        header_labels: tuple[str, ...],
        start_col_letter: str,
        source_col_letter: str,
    ) -> None:
        """Update a simple summary table with unique values and counts.

        Args:
            ws: Worksheet object.
            unique_values: List of values to summarize.
            header_labels: Headers for the summary table.
            start_col_letter: Column where the summary starts.
            source_col_letter: Column to count values from.
        """
        col_idx = column_index_from_string(start_col_letter)
        value_col = get_column_letter(col_idx)
        count_col = get_column_letter(col_idx + 1)

        ws[f"{value_col}1"] = header_labels[0]
        ws[f"{count_col}1"] = header_labels[1]

        self._clear_column_range(ws, value_col, 2, ws.max_row)
        self._clear_column_range(ws, count_col, 2, ws.max_row)

        for i, val in enumerate(unique_values):
            row_idx = i + 2
            ws[f"{value_col}{row_idx}"].value = str(val).upper() if isinstance(val, str) else val
            ws[f"{count_col}{row_idx}"] = f'=COUNTIF({source_col_letter}2:{source_col_letter}1000, {value_col}{row_idx})'

    def _update_metric_table(self, ws: Worksheet) -> None:
        """Update the metric summary table with formulas and ratio.

        Args:
            ws: Worksheet object.
        """
        metrics = ["People", "Gross PLN", "Ratio", "Net PLN"]
        formulas = [
            "=COUNTA(A2:A1000)",
            "=SUM(F2:F1000)",
            self.ratio,
            "=M2*M3"
        ]

        for i, (label, formula) in enumerate(zip(metrics, formulas)):
            row_idx = i + 1
            ws[f"L{row_idx}"] = label
            ws[f"M{row_idx}"].value = formula

        self.set_column_format("M", "0", 2, 2)
        self.set_column_format("M", "0.00", 3, 3)
        self.set_column_format("M", "0", 4, 4)

    def _validate_headers(self) -> None:
        """Validate that headers for all tables are properly defined.

        Raises:
            ValueError: If headers are missing or invalid.
        """
        if not self.main_table_headers or len(self.main_table_headers) < 1:
            raise ValueError("Main table headers collection should have at least 1 item")

        if len(self.company_table_headers) != 2:
            raise ValueError("Insurance Company table headers collection should have 2 items")

        if len(self.summary_table_headers) != 2:
            raise ValueError("Summary table headers collection should have 2 items")

    def _validate_column_ranges(self) -> str:
        """Validate that column ranges for tables do not overlap.

        Returns:
            str: The starting column letter for the summary table.

        Raises:
            ValueError: If column ranges overlap.
        """

        def get_range(start_col: str, width: int) -> set[int]:
            start = column_index_from_string(start_col)
            return set(range(start, start + width))

        ranges = {
            "main": get_range(self.main_table_start_col, len(self.main_table_headers)),
            "insurance_company": get_range(self.company_table_start_col, len(self.company_table_headers)),
        }

        all_used: list[int] = []

        for name, r in ranges.items():
            for other_name, other_r in ranges.items():
                if other_name != name and not r.isdisjoint(other_r):
                    raise ValueError(f"Incorrect columns ranges: {name} and {other_name}")
            all_used.extend(r)

        return get_column_letter(max(all_used) + 2)

    def _aplay_uppercase(self) -> None:
        """Apply uppercase transformation to configured columns."""
        for col in self.uppercase_columns:
            last_row = self.get_last_row_in_col(col)
            if last_row >= 2:
                cell_range = f"{col}2:{col}{last_row}"
                self.apply_str_conversion_for_ranges(lambda v: v.upper(), [cell_range])

    def _style_summary_table(self) -> None:
        """Apply styling to the summary metrics table."""
        if not self.header_style and not self.row_style:
            return

        start_col_idx = column_index_from_string(self.summary_table_start_col)
        label_col_letter = get_column_letter(start_col_idx)
        value_col_letter = get_column_letter(start_col_idx + 1)

        for row_idx in range(1, 5):
            if self.header_style:
                self.style_cell(f"{label_col_letter}{row_idx}", self.header_style)

            if self.row_style:
                self.style_cell(f"{value_col_letter}{row_idx}", self.row_style)
