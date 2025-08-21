from unittest.mock import patch

import pytest
from openpyxl.styles import Font, PatternFill, Alignment, Border

from src.excel.manager.base_manager import ExcelManager
from pathlib import Path

from src.excel.type.style_type import CellStyle, apply_style


def fill_style() -> CellStyle:
    return {"fill": PatternFill("solid")}

def alignment_style() -> CellStyle:
    return {"alignment": Alignment("center")}

def bold_font_style() -> CellStyle:
    return {"font": Font(bold=True)}

def italic_font_style() -> CellStyle:
    return {"font": Font(italic=True)}

def name_font_style() -> CellStyle:
    return {"font": Font(name="Arial")}

def test_add_sheet(example_base_manager: ExcelManager) -> None:
    sheet_name = "test"
    sheet = example_base_manager.get_sheet(sheet_name)

    assert sheet.title == "test"

def test_save(tmp_path: Path) -> None:
    tmp_file = tmp_path / "test.xlsx"
    example_base_manager = ExcelManager(str(tmp_file))
    example_base_manager.save()

    assert tmp_file.exists()

def test_add_row(example_base_manager: ExcelManager) -> None:
    sheet_name = "test"
    example_base_manager.add_row(sheet_name, data={"A": 1})

    sheet = example_base_manager.get_sheet(sheet_name)

    assert sheet["A1"].value == 1


def test_aplay_uppercase(example_base_manager: ExcelManager) -> None:
    sheet_name = "test"
    example_base_manager.add_row(sheet_name, data={"A": "Hello", "B": "World"})
    sheet = example_base_manager.get_sheet(sheet_name)

    def to_upper(s: str) -> str:
        return s.upper()

    example_base_manager.sheet_name = sheet_name
    example_base_manager.apply_str_conversion_for_ranges(
        converter_fn=to_upper,
        uppercase_cell_ranges=["A1:B1"]
    )

    assert sheet["A1"].value == "HELLO"
    assert sheet["B1"].value == "WORLD"

def test_style_cell(example_base_manager: ExcelManager) -> None:
    sheet_name = "test"
    example_base_manager.add_row(sheet_name, data={"A": "Hello", "B": "World"})

    sheet = example_base_manager.get_sheet(sheet_name)
    example_base_manager.style_cell("A1", style=fill_style(), sheet_name=sheet_name)
    example_base_manager.style_cell("B1", style=name_font_style(), sheet_name=sheet_name)

    assert sheet["A1"].fill.fill_type == "solid"
    assert sheet["B1"].font.name == "Arial"

def test_style_table_area(example_base_manager: ExcelManager) -> None:
    sheet_name = "test"
    example_base_manager.add_row(sheet_name, data={"A": "Name", "B": "Email"}, row_idx=1, col_letter="A")
    example_base_manager.add_row(sheet_name, data={"A": "Another", "B": "Row"}, row_idx=2, col_letter="A")

    sheet = example_base_manager.get_sheet(sheet_name)

    example_base_manager.style_table_area(
        "A", ["Name", "Email"], header_style=alignment_style(), row_style=bold_font_style(), sheet_name=sheet_name)

    cell_a = sheet["A1"]
    cell_b = sheet["B1"]
    assert cell_a.value == "Name"
    assert cell_b.value == "Email"
    assert sheet["A1"].alignment.horizontal == "center"
    assert sheet["A2"].font.bold == True

def test_set_column_format(example_base_manager: ExcelManager) -> None:
    sheet_name = "test"
    example_base_manager.add_row(sheet_name, data={"A": "Hello", "B": "World"})
    sheet = example_base_manager.get_sheet(sheet_name)
    example_base_manager.set_column_format("A", "0.00", sheet_name=sheet_name)

    assert sheet["A1"].number_format == "0.00"


def test_get_last_row_in_col_is_empty(example_base_manager: ExcelManager) -> None:
    sheet_name = "test"
    example_base_manager.add_row(sheet_name, data={"A": "Hello"},row_idx=1, col_letter="A")
    example_base_manager.add_row(sheet_name, data={"A": "Another"},row_idx=2, col_letter="A")

    last_row = example_base_manager.get_last_row_in_col("A", sheet_name=sheet_name)
    assert last_row == 2

    last_row_empty = example_base_manager.get_last_row_in_col("C", sheet_name=sheet_name)
    assert last_row_empty == 1


def test_autofit_column_withs_sets_correct_withs(example_base_manager: ExcelManager) -> None:
    sheet_name = 'test'
    example_base_manager.add_row(sheet_name, data={"A": "MuchLongerValueHere", "B": "test_b"})

    example_base_manager.autofit_column_widths(sheet_name, offset_dim=2)
    sheet = example_base_manager.get_sheet(sheet_name)
    width = sheet.column_dimensions["A"].width

    assert width >= len("MuchLongerValueHere") + 2

def test_autofit_column_withs_exception_is_caught(example_base_manager: ExcelManager) -> None:
    sheet_name = "test"
    example_base_manager.add_row(sheet_name, data={"A": "Hello", "B": "World"})

    with patch("builtins.len", side_effect=RuntimeError("Err")):
        try:
            example_base_manager.autofit_column_widths(sheet_name, offset_dim=2)
        except RuntimeError:
            pytest.fail('Exception inside autofit_column_widths() was not caught')


def test_load_or_create(tmp_path: Path) -> None:
    tmp_file = tmp_path / "test.xlsx"
    sheet_name = "test"
    example_base_manager1 = ExcelManager(str(tmp_file))
    example_base_manager1.workbook.create_sheet(sheet_name)
    example_base_manager1.add_row(sheet_name, data={"A": "Hello", "B": "World"})
    example_base_manager1.save()

    example_base_manager2 = ExcelManager(str(tmp_file))
    example_base_manager2._load_or_create()
    sheet = example_base_manager2.get_sheet(sheet_name)

    assert sheet["A1"].value == "Hello"


def test_style_border_side(example_base_manager: ExcelManager) -> None:
    sheet_name = "test"
    sheet = example_base_manager.get_sheet(sheet_name)

    cell = sheet["A1"]

    style: CellStyle = {
        "border_sides": {
            "top": {"style": "thin", "color": "FF0000"},
            "bottom": {"style": "thick", "color": "00FF0000"},
            "left": {"style": "thin", "color": "00FF0000"},
            "right": {"style": "thin", "color": "00FF0000"},
        }
    }

    apply_style(cell, style)

    border: Border = cell.border

    assert border.top is not None
    assert border.bottom.style == "thick"
    assert border.left.color.rgb == "00FF0000"
    assert border.right.style == "thin"









