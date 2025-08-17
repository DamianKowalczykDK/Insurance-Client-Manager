from src.excel.manager.client_manager import ClientExcelManager
from openpyxl.styles import Font, Alignment, PatternFill
from src.service.invoice_service import InvoiceService
from src.service.client_service import ClientService
from src.service.email_service import EmailService
from src.excel.type.style_type import CellStyle
from dotenv import load_dotenv
import os

load_dotenv()


header_style: CellStyle = {
        "font": Font(bold=True, color="000000"),
        "fill": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border_sides": {
            "top": {"style": "medium", "color": "000000"},
            "bottom": {"style": "medium", "color": "000000"},
            "left": {"style": "medium", "color": "000000"},
            "right": {"style": "medium", "color": "000000"},
        }
    }

row_style: CellStyle = {
    "font": Font(name="Calibri", size=11, color="000000"),
    "fill": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
    "alignment": Alignment(horizontal="left", vertical="center"),
    "border_sides": {
        "top": {"style": "thin", "color": "000000"},
        "bottom": {"style": "thin", "color": "000000"},
        "left": {"style": "thin", "color": "000000"},
        "right": {"style": "thin", "color": "000000"},
    }
}

overdue_style: CellStyle = {
    "fill": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "font": Font(color="000000")
}
client_excel_manager = ClientExcelManager(
    filepath="Clients.xlsx",
    sheet_name="Clients",
    main_table_headers=["NAME", "EMAIL", "INSURANCE_COMPANY", "CAR_MODEL", "CAR_YEAR", "PRICE", "NEXT_PAYMENT"],

    header_style=header_style,
    row_style=row_style,
    overdue_style=overdue_style,

    main_table_start_col="A",
    company_table_start_col="I",

)

smtp_server = os.getenv("SMTP_SERVER")
port = int(os.getenv("SMTP_PORT"))
sender_email = os.getenv("SENDER_EMAIL")
sender_password = os.getenv("SENDER_PASSWORD")
email_service = EmailService(smtp_server, port, sender_email, sender_password)

invoice_service = InvoiceService(
    api_token=os.getenv("INVOICE_API_TOKEN"),
    domain=os.getenv("INVOICE_DOMAIN"),
)

client_service = ClientService(client_excel_manager, email_service, invoice_service)


def create_client_service() -> ClientService:
    # ------------------------------------------------------------------------------------------------------------------
    # STUDENT EXCEL MANAGER
    # ------------------------------------------------------------------------------------------------------------------
    header_style: CellStyle = {
        "font": Font(bold=True, color="000000"),
        "fill": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border_sides": {
            "top": {"style": "medium", "color": "000000"},
            "bottom": {"style": "medium", "color": "000000"},
            "left": {"style": "medium", "color": "000000"},
            "right": {"style": "medium", "color": "000000"},
        }
    }

    row_style: CellStyle = {
        "font": Font(name="Calibri", size=11, color="000000"),
        "fill": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border_sides": {
            "top": {"style": "thin", "color": "000000"},
            "bottom": {"style": "thin", "color": "000000"},
            "left": {"style": "thin", "color": "000000"},
            "right": {"style": "thin", "color": "000000"},
        }
    }

    overdue_style: CellStyle = {
        "fill": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
        "font": Font(color="000000")
    }
    client_excel_manager = ClientExcelManager(
        filepath="Clients.xlsx",
        sheet_name="Clients",
        main_table_headers=["NAME", "EMAIL", "INSURANCE_COMPANY", "CAR_MODEL", "CAR_YEAR", "PRICE", "NEXT_PAYMENT"],

        header_style=header_style,
        row_style=row_style,
        overdue_style=overdue_style,

        main_table_start_col="A",
        company_table_start_col="I",

    )

    smtp_server = os.getenv("SMTP_SERVER")
    port = int(os.getenv("SMTP_PORT"))
    sender_email = os.getenv("SENDER_EMAIL")
    sender_password = os.getenv("SENDER_PASSWORD")
    email_service = EmailService(smtp_server, port, sender_email, sender_password)

    invoice_service = InvoiceService(
        api_token=os.getenv("INVOICE_API_TOKEN"),
        domain=os.getenv("INVOICE_DOMAIN"),
    )

    return ClientService(client_excel_manager, email_service, invoice_service)